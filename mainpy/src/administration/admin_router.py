from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference, LineChart, PieChart
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from collections import defaultdict
import pandas as pd
from io import BytesIO
from mainpy.src.users.user import User, get_user_by_id, get_current_active_user_model
from mainpy.src.services.administration_service import get_users_each_month, get_users_authors, get_user_and_authors, \
    get_posts_by_month, get_posts_each_tag

admin_router = APIRouter()

def is_admin(user_id):
    user = get_user_by_id(user_id)
    if user.role == 'Admin':
        return True
    else:
        return False

def create_excel_chart(data) -> BytesIO:
    wb = Workbook()
    ws = wb.active

    ws.append(['Месяц', 'Пользователи'])
    for elem in data:
        ws.append([elem['month'], elem['users']])


    chart = BarChart()
    chart.title = "Регистрации пользователей"
    chart.x_axis.title = "Месяц"
    chart.y_axis.title = "Количество пользователей"

    main_data = Reference(ws, min_col=2, min_row=1, max_row=len(data)+1)
    categories = Reference(ws, min_col=1, min_row=2, max_row=len(data)+1)


    chart.add_data(main_data, titles_from_data=True)
    chart.set_categories(categories)

    chart.y_axis.min = 2
    chart.y_axis.max = max(entry['users'] for entry in data) + 5
    chart.y_axis.major_tick_mark = "out"

    ws.add_chart(chart, "E5")

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output

def create_excel_active_users(data) -> BytesIO:
    wb = Workbook()
    ws = wb.active

    ws.append(['Месяц', 'Активные пользователи'])
    for elem in data:
        ws.append([elem['month'], elem['users']])


    chart = PieChart()

    main_data = Reference(ws, min_col=2, min_row=1, max_row=len(data)+1)
    categories = Reference(ws, min_col=1, min_row=2, max_row=len(data)+1)


    chart.add_data(main_data, titles_from_data=True)
    chart.set_categories(categories)

    ws.add_chart(chart, "E5")

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output

def create_excel_authors_percentage(data) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "test"

    ws.append(['Месяц', 'Активные пользователи', 'Авторы'])

    for month_data in data:
        month = month_data['month']
        active_users = month_data['users']
        authors = month_data['authors']
        active_user_percentage = 100
        authors_percentage = (authors/active_users)*100

        ws.append([month, active_users, authors, active_user_percentage-authors_percentage, authors_percentage])

    for i in range(len(data)):
        chart = PieChart()
        chart.title = f"Процент авторов за {data[i]['month']}"

        data_ref = Reference(ws, min_col=4, min_row=i + 2,max_row=i+2, max_col=5)
        categories_ref = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=1)

        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(categories_ref)

        chart.width = 17

        ws.add_chart(chart, f"E{(i * 15) + 5}")


    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output

def create_excel_for_posts(data):
    wb = Workbook()
    ws = wb.active

    ws.append(['Месяц', 'Публикации'])
    for elem in data:
        ws.append([elem['month'], elem['posts']])

    chart = BarChart()
    chart.title = "Публикации по месяцам"

    main_data = Reference(ws, min_col=2, min_row=1, max_row=len(data) + 1)
    categories = Reference(ws, min_col=1, min_row=2, max_row=len(data) + 1)

    chart.add_data(main_data, titles_from_data=True)
    chart.set_categories(categories)

    ws.add_chart(chart, "E5")

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output

def create_excel_for_posts_tags(data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Лист 1"

    ws.append(['Месяц', 'Количество постов', 'Тема'])

    tags_count = []
    values_count = []

    for month_data in data:
        temp = 0
        for i in range(len(month_data['array_count'])):
            temp += month_data['array_count'][i]
        tags_count.append(temp)
        values_count.append(len(month_data['array_count']))

    for i in range(len(data)):
        ws.append([data[i]['month'],data[i]['array_count'][0],data[i]['array_tags'][0], (data[i]['array_count'][0]/tags_count[i])*100])
        for j in range(1,len(data[i]['array_count'])):
            ws.append([None,data[i]['array_count'][j],data[i]['array_tags'][j],(data[i]['array_count'][j]/tags_count[i])*100])
    amount = 0
    for i in range(len(data)):
        chart = PieChart()
        chart.title = f"Процент тегов постов за {data[i]['month']}"
        data_ref = Reference(ws, min_col=4, min_row=amount + 2-1,max_row=amount+values_count[i]+1, max_col=4)
        categories_ref = Reference(ws, min_col=3, min_row=amount+2, max_col=3, max_row=amount+1+values_count[i])
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(categories_ref)
        amount += values_count[i]

        chart.width = 17

        ws.add_chart(chart, f"E{(i * 15) + 5}")

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output

@admin_router.get('/users/csv')
async def get_users_csv():
    data = get_users_each_month()
    return data

@admin_router.get('/users_each_month')
async def get_users_by_month(current_user = Depends(get_current_active_user_model)):
    if is_admin(current_user.id):
        user_data = get_users_each_month()
        excel_file = create_excel_chart(user_data)
        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=registrations.xlsx"}
        )
    else:
        raise HTTPException(status_code=403, detail="Not Authorized")

@admin_router.get('/users_authors')
async def get_users_are_authors(current_user = Depends(get_current_active_user_model)):
    if is_admin(current_user.id):
        user_data = get_users_authors()
        excel_file = create_excel_active_users(user_data)
        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=active_users.xlsx"}
        )
    else:
        raise HTTPException(status_code=403, detail="Not Authorized")

@admin_router.get('/users_authors_percentage')
async def get_users_authors_percentage(current_user = Depends(get_current_active_user_model)):
    if is_admin(current_user.id):
        data = get_user_and_authors()
        excel_file = create_excel_authors_percentage(data)
        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=users_percentage.xlsx"}
        )
    else:
        raise HTTPException(status_code=403, detail="Not Authorized")

@admin_router.get('/posts_by_month')
async def get_posts_each_month(current_user = Depends(get_current_active_user_model)):
    if is_admin(current_user.id):
        data = get_posts_by_month()
        excel_file = create_excel_for_posts(data)
        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=posts_by_month.xlsx"}
        )
    else:
        raise HTTPException(status_code=403, detail="Not Authorized")

@admin_router.get('/posts_by_tags')
async def get_posts_by_tags(current_user = Depends(get_current_active_user_model)):
    if is_admin(current_user.id):
        data = get_posts_each_tag()
        excel_file = create_excel_for_posts_tags(data)
        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=posts_by_tags_month.xlsx"}
        )
    else:
        raise HTTPException(status_code=403, detail="Not Authorized")